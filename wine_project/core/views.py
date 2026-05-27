import io
import os
import urllib.request
import datetime
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab Imports for beautiful PDF exports
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def get_thai_font():
    """
    Downloads the Google Font 'Sarabun' (Regular & Bold) if not locally present,
    and registers it with ReportLab. Falls back to macOS system 'Ayuthaya' or 'Helvetica'.
    """
    base_dir = os.path.dirname(__file__)
    font_dir = os.path.join(base_dir, 'static', 'core', 'fonts')
    os.makedirs(font_dir, exist_ok=True)
    
    reg_path = os.path.join(font_dir, 'Sarabun-Regular.ttf')
    bold_path = os.path.join(font_dir, 'Sarabun-Bold.ttf')
    
    # Try downloading the Sarabun font files from Google Fonts repository if missing
    if not os.path.exists(reg_path) or not os.path.exists(bold_path):
        try:
            req_reg = urllib.request.Request(
                'https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Regular.ttf',
                headers={'User-Agent': 'Mozilla/5.0'}
            )
            req_bold = urllib.request.Request(
                'https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Bold.ttf',
                headers={'User-Agent': 'Mozilla/5.0'}
            )
            with urllib.request.urlopen(req_reg) as response, open(reg_path, 'wb') as out_file:
                out_file.write(response.read())
            with urllib.request.urlopen(req_bold) as response, open(bold_path, 'wb') as out_file:
                out_file.write(response.read())
        except Exception as e:
            print(f"Error downloading Google Fonts: {e}")
            
    if os.path.exists(reg_path) and os.path.exists(bold_path):
        try:
            pdfmetrics.registerFont(TTFont('Sarabun', reg_path))
            pdfmetrics.registerFont(TTFont('Sarabun-Bold', bold_path))
            return 'Sarabun', 'Sarabun-Bold'
        except Exception as e:
            print(f"Error registering Sarabun fonts: {e}")
            
    # macOS system font fallback
    mac_font = '/System/Library/Fonts/Supplemental/Ayuthaya.ttf'
    if os.path.exists(mac_font):
        try:
            pdfmetrics.registerFont(TTFont('Ayuthaya', mac_font))
            return 'Ayuthaya', 'Ayuthaya'
        except Exception as e:
            print(f"Error registering Ayuthaya font fallback: {e}")
            
    # Standard Helvetica default
    return 'Helvetica', 'Helvetica-Bold'

def clean_wine_dataframe(file_obj):
    # อ่าน Excel จาก Memory (ไม่ต้องเซฟลงเครื่อง)
    df_base = pd.read_excel(file_obj, header=[4, 5])
    
    new_columns = []
    for col in df_base.columns:
        top_header = str(col[0]).strip()
        bottom_header = str(col[1]).strip()
        if top_header.startswith("Unnamed:") or top_header == "nan": col_name = bottom_header
        elif bottom_header.startswith("Unnamed:") or bottom_header == "nan": col_name = top_header
        else: col_name = f"{top_header}|{bottom_header}"
        new_columns.append(col_name)
    df_base.columns = [c.strip() for c in new_columns]

    def find_col(keyword):
        for c in df_base.columns:
            if keyword in c: return c
        return keyword

    size_col = find_col('ขนาด')
    if size_col in df_base.columns:
        df_base[size_col] = df_base[size_col].astype(str).str.replace(r'\(Wooden Gift Box\)', '', case=False, regex=True).str.strip()

    sort_col_1 = find_col('หมวดหมู่รอง')
    sort_col_2 = find_col('หมวดหมู่หลัก')
    return df_base.sort_values(by=[sort_col_1, sort_col_2]).reset_index(drop=True), find_col

def prepare_sheet_data(df_source, columns_mapping, find_col_fn, is_private=False, is_private_1=False, is_horeca=False):
    """
    Cleans and filters data according to the sheet type.
    """
    df_sheet = df_source.copy()
    sku_col = find_col_fn('รหัสสินค้า')
    qty_col = find_col_fn('ทั้งหมด')
    subcat_col = find_col_fn('หมวดหมู่รอง')
    private_price_col = find_col_fn('Private')
    horeca_price_col = find_col_fn('HORECA')
    
    if is_private or is_private_1:
        if private_price_col in df_sheet.columns:
            df_sheet[private_price_col] = pd.to_numeric(df_sheet[private_price_col], errors='coerce').fillna(0)
            df_sheet = df_sheet[df_sheet[private_price_col] != 0]
    if is_horeca:
        if horeca_price_col in df_sheet.columns:
            df_sheet[horeca_price_col] = pd.to_numeric(df_sheet[horeca_price_col], errors='coerce').fillna(0)
            df_sheet = df_sheet[df_sheet[horeca_price_col] != 0]

    if is_private or is_private_1:
        df_sheet.loc[df_sheet[sku_col].isin(['Vc001', 'Mcd001']), qty_col] = 300
        df_sheet = df_sheet[~df_sheet[sku_col].astype(str).str.contains('WD', na=False)]
    if is_private_1:
        df_sheet = df_sheet[~df_sheet[subcat_col].isin(['Australia', 'Germany', 'New Zealand'])]
        
    return df_sheet

def serialize_sheet_data(df_sheet, columns_mapping, find_col_fn):
    """
    Serializes a sheet dataframe into a JSON structure for dynamic client-side previews.
    """
    if df_sheet.empty:
        return None
        
    sku_col = find_col_fn('รหัสสินค้า')
    qty_col = find_col_fn('ทั้งหมด')
    subcat_col = find_col_fn('หมวดหมู่รอง')
    
    headers = ['No'] + list(columns_mapping.values())
    grouped = df_sheet.groupby(subcat_col, sort=False)
    
    groups_list = []
    global_no = 1
    
    for country, group in grouped:
        if group.empty or pd.isna(country):
            continue
            
        rows_list = []
        for _, row in group.iterrows():
            row_dict = {}
            row_dict['No'] = global_no
            
            for orig_col, mapped_col in columns_mapping.items():
                val = row[orig_col]
                if pd.isna(val):
                    val = ""
                elif isinstance(val, (int, float)):
                    # Pass numbers raw so JS can format them correctly
                    val = val
                else:
                    val = str(val).strip()
                row_dict[mapped_col] = val
                
            rows_list.append(row_dict)
            global_no += 1
            
        groups_list.append({
            "country": str(country).upper(),
            "rows": rows_list
        })
        
    return {
        "headers": headers,
        "groups": groups_list
    }

def write_sheet_with_styling(wb, sheet_name, df_sheet, columns_mapping, find_col_fn):
    if df_sheet.empty: return

    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True  
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT   
    ws.page_setup.paperSize = ws.PAPERSIZE_A4             
    ws.page_margins.left, ws.page_margins.right = 0.3, 0.3
    ws.page_margins.top, ws.page_margins.bottom = 0.5, 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True      
    ws.page_setup.fitToWidth = 1                          
    ws.page_setup.fitToHeight = 0                         
    ws.print_title_rows = '1:2'  # ซ้ำหัวกระดาษทุกหน้าเวลาพิมพ์ PDF

    font_name = 'Cordia New'
    ws.row_dimensions[1].height = 30  
    
    title_cell = ws.cell(row=1, column=1, value="DEEPLUS JUD HAI Co., Ltd.")
    title_cell.font = Font(name=font_name, size=18, bold=True, color='5C0614')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    
    date_cell = ws.cell(row=1, column=6, value="=TODAY()")
    date_cell.font = Font(name=font_name, size=14, bold=False, color='555555')
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    date_cell.number_format = 'yyyy-mm-dd'
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
    
    ws.freeze_panes = 'A4'
    
    header_font = Font(name=font_name, size=15, bold=True, color='FFFFFF')
    country_font = Font(name=font_name, size=14, bold=True, color='5C0614')
    body_font = Font(name=font_name, size=13, bold=False, color='000000')
    header_fill = PatternFill(start_color='5C0614', end_color='5C0614', fill_type='solid')
    country_fill = PatternFill(start_color='FDF4F5', end_color='FDF4F5', fill_type='solid')
    thin_side = Side(style='thin', color='D3D3D3')
    dashed_side = Side(style='dashed', color='A0A0A0')
    
    headers = ['No'] + list(columns_mapping.values())
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(style='medium', color='000000'))
    ws.row_dimensions[2].height = 28
    
    wine_desc_col_idx = 3  
    if 'Wine Description' in headers: wine_desc_col_idx = headers.index('Wine Description') + 1
    
    subcat_col = find_col_fn('หมวดหมู่รอง')
    grouped = df_sheet.groupby(subcat_col, sort=False)
    global_no, current_row = 1, 3
    
    for country, group in grouped:
        if group.empty or pd.isna(country): continue
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.fill = country_fill
            c.border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        country_cell = ws.cell(row=current_row, column=wine_desc_col_idx, value=f"--- {str(country).upper()} ---")
        country_cell.font = country_font
        country_cell.alignment = Alignment(horizontal='center', vertical='center') 
        ws.row_dimensions[current_row].height = 24  
        current_row += 1
        
        num_rows = len(group)
        for idx, (_, row) in enumerate(group.iterrows()):
            ws.cell(row=current_row, column=1, value=global_no)
            for c_idx, orig_col in enumerate(columns_mapping.keys()):
                ws.cell(row=current_row, column=c_idx + 2, value=row[orig_col])
                
            for col_idx, col_name in enumerate(headers):
                cell = ws.cell(row=current_row, column=col_idx + 1)
                cell.font = body_font
                if col_name in ['No', 'SKU', 'หมวดหมู่รอง', 'QTY']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_name in ['Retail Price', 'Private', 'ต้นทุน', 'HORECA', 'Ltr.']:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if col_name == 'Ltr.':
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0.00'
                        except: cell.alignment = Alignment(horizontal='center', vertical='center')
                    else: cell.number_format = '#,##0'
                else: cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if idx < num_rows - 1: cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=dashed_side)
                else: cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            ws.row_dimensions[current_row].height = 20
            global_no += 1
            current_row += 1

    thai_non_spacing = set([ord(c) for c in 'ิีึืุู้่๊๋็์ัํ'])
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith('='): approx_width = 12
                else:
                    approx_width = 0
                    for char in val_str:
                        if ord(char) in thai_non_spacing: continue  
                        if ord(char) > 127 or char.isupper(): approx_width += 1.2
                        else: approx_width += 0.95
                if approx_width > max_len: max_len = approx_width
        if col_letter == 'C': ws.column_dimensions[col_letter].width = max(max_len + 6, 35)
        else: ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

def preview_file(request):
    """
    Endpoint that receives the Excel file and extracts cleaned data for live AJAX preview.
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            df_base_sorted, find_col = clean_wine_dataframe(excel_file)
            
            retail_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Retail Price'): 'Retail Price'}
            private_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('หมวดหมู่หลัก'): 'หมวดหมู่หลัก', find_col('หมวดหมู่รอง'): 'หมวดหมู่รอง', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Private'): 'Private', find_col('รหัสอ้างอิงสินค้า2'): 'Point'}
            cost_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('ต้นทุน 1 เฉลี่ย'): 'ต้นทุน'}
            horeca_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('HORECA'): 'HORECA'}

            preview_data = {
                "retail": serialize_sheet_data(prepare_sheet_data(df_base_sorted, retail_mapping, find_col), retail_mapping, find_col),
                "private": serialize_sheet_data(prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private=True), private_mapping, find_col),
                "cost": serialize_sheet_data(prepare_sheet_data(df_base_sorted, cost_mapping, find_col), cost_mapping, find_col),
                "horeca": serialize_sheet_data(prepare_sheet_data(df_base_sorted, horeca_mapping, find_col, is_horeca=True), horeca_mapping, find_col),
                "private_1": serialize_sheet_data(prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private_1=True), private_mapping, find_col),
            }
            
            return JsonResponse({"success": True, "preview": preview_data, "filename": excel_file.name})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
            
    return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

def export_pdf(request):
    """
    Generates and returns a beautifully structured PDF document with selected sheets.
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            selected_sheets = request.POST.getlist('sheets')
            
            if not selected_sheets:
                return HttpResponse("Please select at least one sheet to export.", status=400)
                
            df_base_sorted, find_col = clean_wine_dataframe(excel_file)
            
            retail_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Retail Price'): 'Retail Price'}
            private_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('หมวดหมู่หลัก'): 'หมวดหมู่หลัก', find_col('หมวดหมู่รอง'): 'หมวดหมู่รอง', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Private'): 'Private', find_col('รหัสอ้างอิงสินค้า2'): 'Point'}
            cost_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('ต้นทุน 1 เฉลี่ย'): 'ต้นทุน'}
            horeca_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('HORECA'): 'HORECA'}

            buffer = io.BytesIO()
            
            # Setup A4 document with 0.3 inches (21.6 pt) left/right margins and 0.5 inches (36 pt) top/bottom margins
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                leftMargin=21.6,
                rightMargin=21.6,
                topMargin=36,
                bottomMargin=36
            )
            
            story = []
            
            # Autodetect or download the Thai Sarabun Font
            font_regular, font_bold = get_thai_font()
            
            styles = getSampleStyleSheet()
            styles['Normal'].fontName = font_regular
            
            # Styling definitions
            title_style = ParagraphStyle(
                'CompanyTitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=15,
                textColor=colors.HexColor('#5C0614'),
                spaceAfter=2
            )
            
            subtitle_style = ParagraphStyle(
                'SheetSubtitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=11,
                textColor=colors.HexColor('#444444'),
                spaceAfter=12
            )
            
            header_cell_style = ParagraphStyle(
                'HeaderCell',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=9,
                textColor=colors.white,
                alignment=1
            )
            
            country_cell_style = ParagraphStyle(
                'CountryCell',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=9,
                textColor=colors.HexColor('#5C0614'),
                alignment=1
            )
            
            body_cell_left = ParagraphStyle(
                'BodyLeft',
                parent=styles['Normal'],
                fontName=font_regular,
                fontSize=8.5,
                textColor=colors.black,
                alignment=0
            )
            
            body_cell_center = ParagraphStyle(
                'BodyCenter',
                parent=styles['Normal'],
                fontName=font_regular,
                fontSize=8.5,
                textColor=colors.black,
                alignment=1
            )
            
            body_cell_right = ParagraphStyle(
                'BodyRight',
                parent=styles['Normal'],
                fontName=font_regular,
                fontSize=8.5,
                textColor=colors.black,
                alignment=2
            )
            
            sheets_to_process = []
            if 'retail' in selected_sheets:
                sheets_to_process.append(('Retail Price', prepare_sheet_data(df_base_sorted, retail_mapping, find_col), retail_mapping))
            if 'private' in selected_sheets:
                sheets_to_process.append(('Private', prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private=True), private_mapping))
            if 'cost' in selected_sheets:
                sheets_to_process.append(('ต้นทุน', prepare_sheet_data(df_base_sorted, cost_mapping, find_col), cost_mapping))
            if 'horeca' in selected_sheets:
                sheets_to_process.append(('HORECA', prepare_sheet_data(df_base_sorted, horeca_mapping, find_col, is_horeca=True), horeca_mapping))
            if 'private_1' in selected_sheets:
                sheets_to_process.append(('Private-1', prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private_1=True), private_mapping))

            first_sheet = True
            for sheet_title, df_sheet, columns_mapping in sheets_to_process:
                if df_sheet.empty:
                    continue
                    
                if not first_sheet:
                    story.append(PageBreak())
                first_sheet = False
                
                # Header Section
                story.append(Paragraph("DEEPLUS JUD HAI Co., Ltd.", title_style))
                today_str = datetime.date.today().strftime('%Y-%m-%d')
                story.append(Paragraph(f" {sheet_title} | วันที่พิมพ์: {today_str}", subtitle_style))
                
                # Build Columns Structure
                headers = ['No'] + list(columns_mapping.values())
                
                # Width definitions: Total width is 552 points
                if sheet_title in ['Private', 'Private-1']:
                    # 10 columns
                    col_widths = [18, 48, 154, 42, 54, 54, 32, 32, 56, 32]
                else:
                    # 7 columns
                    col_widths = [18, 52, 270, 52, 32, 38, 90]
                    
                table_data = [[Paragraph(h, header_cell_style) for h in headers]]
                row_styles = [] # List of tuples: (row_idx, row_type, extra)
                
                subcat_col = find_col('หมวดหมู่รอง')
                grouped = df_sheet.groupby(subcat_col, sort=False)
                
                global_no = 1
                current_row_idx = 1
                
                for country, group in grouped:
                    if group.empty or pd.isna(country):
                        continue
                    
                    # Country Group Divider Row
                    country_text = f"--- {str(country).upper()} ---"
                    country_row = [Paragraph("", body_cell_center)] * len(headers)
                    country_row[0] = Paragraph(country_text, country_cell_style)
                    table_data.append(country_row)
                    
                    row_styles.append((current_row_idx, 'country', len(headers)))
                    current_row_idx += 1
                    
                    num_rows = len(group)
                    for idx, (_, row) in enumerate(group.iterrows()):
                        row_data = []
                        # Number
                        row_data.append(Paragraph(str(global_no), body_cell_center))
                        
                        for c_idx, orig_col in enumerate(columns_mapping.keys()):
                            mapped_col = headers[c_idx + 1]
                            val = row[orig_col]
                            
                            if pd.isna(val):
                                val_str = ""
                            elif mapped_col in ['Retail Price', 'Private', 'ต้นทุน', 'HORECA']:
                                try:
                                    val_str = "{:,.0f}".format(float(val))
                                except:
                                    val_str = str(val)
                            elif mapped_col == 'Ltr.':
                                try:
                                    val_str = "{:,.2f}".format(float(val))
                                except:
                                    val_str = str(val)
                            else:
                                val_str = str(val)
                                
                            # Alignment Mapping
                            if mapped_col in ['No', 'SKU', 'หมวดหมู่รอง', 'QTY']:
                                align_style = body_cell_center
                            elif mapped_col in ['Retail Price', 'Private', 'ต้นทุน', 'HORECA', 'Ltr.']:
                                align_style = body_cell_right
                            else:
                                align_style = body_cell_left
                                
                            row_data.append(Paragraph(val_str, align_style))
                            
                        table_data.append(row_data)
                        row_styles.append((current_row_idx, 'body', idx == num_rows - 1))
                        global_no += 1
                        current_row_idx += 1
                
                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                t_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5C0614')),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                    ('TOPPADDING', (0, 0), (-1, 0), 4),
                ]
                
                thin_color = colors.HexColor('#E0E0E0')
                dashed_color = colors.HexColor('#C0C0C0')
                
                for idx, r_type, extra in row_styles:
                    if r_type == 'country':
                        t_style.append(('SPAN', (0, idx), (extra - 1, idx)))
                        t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FDF4F5')))
                        t_style.append(('TOPPADDING', (0, idx), (-1, idx), 4))
                        t_style.append(('BOTTOMPADDING', (0, idx), (-1, idx), 4))
                        t_style.append(('LINEBELOW', (0, idx), (-1, idx), 0.5, thin_color))
                        t_style.append(('LINEABOVE', (0, idx), (-1, idx), 0.5, thin_color))
                    elif r_type == 'body':
                        # If bottom cell in country group, solid border, otherwise dashed
                        line_style = thin_color if extra else dashed_color
                        t_style.append(('LINEBELOW', (0, idx), (-1, idx), 0.5, line_style))
                        t_style.append(('LINEBEFORE', (0, idx), (-1, idx), 0.5, thin_color))
                        t_style.append(('LINEAFTER', (0, idx), (-1, idx), 0.5, thin_color))
                        t_style.append(('TOPPADDING', (0, idx), (-1, idx), 3))
                        t_style.append(('BOTTOMPADDING', (0, idx), (-1, idx), 3))
                        
                t.setStyle(TableStyle(t_style))
                story.append(t)
                
            doc.build(story)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="deeplus_wine_report_{datetime.date.today().strftime("%Y%m%d")}.pdf"'
            response.write(pdf_data)
            return response
        except Exception as e:
            return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
            
    return HttpResponse("Invalid Request", status=400)

def upload_file(request):
    """
    Main landing route and Excel Clean & Export endpoint.
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            selected_sheets = request.POST.getlist('sheets')
            
            if not selected_sheets:
                return HttpResponse("Please select at least one sheet to download.", status=400)

            df_base_sorted, find_col = clean_wine_dataframe(excel_file)
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            retail_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Retail Price'): 'Retail Price'}
            private_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('หมวดหมู่หลัก'): 'หมวดหมู่หลัก', find_col('หมวดหมู่รอง'): 'หมวดหมู่รอง', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('Private'): 'Private', find_col('รหัสอ้างอิงสินค้า2'): 'Point'}
            cost_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('ต้นทุน 1 เฉลี่ย'): 'ต้นทุน'}
            horeca_mapping = {find_col('รหัสสินค้า'): 'SKU', find_col('ชื่อสินค้า'): 'Wine Description', find_col('กลุ่มสินค้า'): 'Group', find_col('ขนาด'): 'Ltr.', find_col('ทั้งหมด'): 'QTY', find_col('HORECA'): 'HORECA'}

            if 'retail' in selected_sheets: 
                write_sheet_with_styling(wb, 'Retail Price', prepare_sheet_data(df_base_sorted, retail_mapping, find_col), retail_mapping, find_col)
            if 'private' in selected_sheets: 
                write_sheet_with_styling(wb, 'Private', prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private=True), private_mapping, find_col)
            if 'cost' in selected_sheets: 
                write_sheet_with_styling(wb, 'ต้นทุน', prepare_sheet_data(df_base_sorted, cost_mapping, find_col), cost_mapping, find_col)
            if 'horeca' in selected_sheets: 
                write_sheet_with_styling(wb, 'HORECA', prepare_sheet_data(df_base_sorted, horeca_mapping, find_col, is_horeca=True), horeca_mapping, find_col)
            if 'private_1' in selected_sheets: 
                write_sheet_with_styling(wb, 'Private-1', prepare_sheet_data(df_base_sorted, private_mapping, find_col, is_private_1=True), private_mapping, find_col)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="deeplus_wine_cleaned_{datetime.date.today().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
        except Exception as e:
            return HttpResponse(f"Error cleaning Excel file: {str(e)}", status=500)

    return render(request, 'core/upload.html')